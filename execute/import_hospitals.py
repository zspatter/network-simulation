import logging
import shelve
from os.path import abspath, join

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from requests_html import HTMLSession

from network_simulator.Network import Network
from network_simulator.Node import Node


def set_default_indices():
    """
    Sets default dict for columns (placeholders for indices)
    """
    fields = ('unique id',
              'hospital name',
              'city',
              'state',
              'region')
    columns = {}
    for field in fields:
        columns.setdefault(field, None)

    return columns


def get_column_indices(worksheet, columns):
    """
    Gets index for each column for reading data

    :param Worksheet worksheet: worksheet to read
    :param dict columns: expected columns to store indices
    """
    for x in range(1, 9):
        cell = worksheet[f'{get_column_letter(x)}1'].value
        if cell and cell.lower() in columns:
            columns[cell.lower()] = get_column_letter(x)


def import_nodes(worksheet):
    """
    Imports node from each row and adds it to the network. These nodes
    only contain the following fields: node_id, hospital_name, region,
    city, and state. (adjacency_dict is handled in generate_distance_vector()
    and status is assumed to be True)

    :param Worksheet worksheet: worksheet to read data from
    :return: Network
    """
    network = Network()
    for x in range(2, worksheet.max_row + 1):
        network.add_node(
                node=Node(node_id=int(worksheet[f'{column_indices["unique id"]}{x}'].value),
                          hospital_name=worksheet[f'{column_indices["hospital name"]}{x}'].value,
                          region=int(worksheet[f'{column_indices["region"]}{x}'].value),
                          city=worksheet[f'{column_indices["city"]}{x}'].value,
                          state=worksheet[f'{column_indices["state"]}{x}'].value),
                feedback=False)

    generate_distance_vector(network=network)
    return network


def generate_distance_vector(network):
    """
    Finds weight from any given node to all other given node in the
    network using the city, state, and region parameters

    :param Network network:
    """
    for source, adjacent in node_pair_generator(network=network):
        weight = lookup_weight(source=source,
                               adjacent=adjacent)
        regional_weight = get_adjacent_regional_weight(source=source,
                                                       adjacent=adjacent)

        if weight and regional_weight:
            network.add_edge(node_id1=source.node_id,
                             node_id2=adjacent.node_id,
                             weight=weight,
                             regional_weight=regional_weight,
                             feedback=False)


def lookup_weight(source, adjacent):
    """
    Determines weight between source and adjacent nodes. Only returns
    a value if there is a weight specified between the nodes.

    :param Node source: source node
    :param Node adjacent: destination node
    """
    source_location = (source.city, source.state, source.region)
    adjacent_location = (adjacent.city, adjacent.state, adjacent.region)

    if source_location in distance_matrix \
            and adjacent_location in distance_matrix[source_location]:
        return distance_matrix[source_location][adjacent_location]


def node_pair_generator(network):
    """
    Generator that yields all possible pairings of nodes

    :param Network network: source of nodes
    """
    for source in network.nodes():
        for adjacent_id in set(network.node_ids()) - {source.node_id}:
            adjacent = network.network_dict[adjacent_id]

            yield source, adjacent


def get_adjacent_regional_weight(source, adjacent, weight=None):
    """
    Calculates weight based up on city, state, and region fields for
    both the node and adjacent

    :param Node source: current source node
    :param Node adjacent: current destination node
    :param float weight: default max weight
    :return: weight
    """

    if source.region == adjacent.region:
        if source.city == adjacent.city and source.state == adjacent.state:
            weight = 1
        elif source.state == adjacent.state and source.city != adjacent.city:
            weight = 2
        elif source.state != adjacent.state:
            weight = 3
    # condition for Virginia
    elif source.state == adjacent.state:
        weight = 3
    elif adjacent.region in neighbor_regions[source.region]:
        weight = 4

    return weight


def get_unique_locations(worksheet):
    """
    Returns a list of all unique city/state/region combinations for hospitals

    :param Worksheet worksheet: worksheet containing information about
                city/state/region of hospitals in network
    """
    # creates a set of unique locations
    locations = set()
    for x in range(2, worksheet.max_row + 1):
        locations.add((worksheet.cell(row=x, column=column_indices['city']).value,
                       worksheet.cell(row=x, column=column_indices['state']).value,
                       int(worksheet.cell(row=x, column=column_indices['region']).value)))

    return sort_locations(list(locations))


def sort_locations(locations):
    """
    Sorts locations collection by region -> state --> city in ascending
    order. This ensures the consistent results with successive executions
    by ordering the collection.

    :param list locations: list of unqiue (city, state, region) tuples
    """
    # sorts collection by region/state/city in ascending order
    sorted_locations = sorted(locations, key=lambda tup: tup[0])
    sorted_locations = sorted(sorted_locations, key=lambda tup: tup[1])
    sorted_locations = sorted(sorted_locations, key=lambda tup: tup[2])

    return sorted_locations


def set_default_distances(locations):
    """
    Sets default distances in the distance matrix (distance_dict)

    :param locations: list of unique locations
    """
    distance_dict = dict()
    for location in locations:
        _, _, origin_region = location
        adjacents = dict()
        for destination_city, destination_state, region in locations:
            if region == origin_region or region in neighbor_regions[origin_region]:
                adjacents.setdefault((destination_city, destination_state, region), None)
        distance_dict[location] = adjacents
    return distance_dict


def get_distances(distance_vector):
    """
    Finds the distance between all connected nodes

    :param dict distance_vector: distance matrix with default weights
    """
    for source in sort_locations(distance_vector.keys()):
        source_city, source_state, source_region = source
        for destination in distance_vector[source]:
            destination_city, destination_state, destination_region = destination
            # if city/state are same for source and destination
            if source == destination:
                distance_vector[source][destination] = 1
                distance_vector[destination][source] = 1
            # verify source/destination are adjacent regions and distance
            # values are not already set
            elif destination_region in neighbor_regions[source_region] \
                    and not (distance_vector[source][destination]
                             or distance_vector[destination][source]):
                distance = get_distance(source_city=source_city,
                                        source_state=source_state,
                                        destination_city=destination_city,
                                        destination_state=destination_state)
                # verify distance value
                if distance:
                    print(f"{f'{source_city}, {source_state}':<30}"
                          f"{f'{destination_city}, {destination_state}':<30}"
                          f"{f'{distance:,.2f} km':>12}")
                    distance_vector[source][destination] = distance
                    distance_vector[destination][source] = distance

    return distance_vector


def get_distance(source_city, source_state, destination_city, destination_state):
    """
    Gathers the distance in km between the source and destination locations

    :param str source_city: name of source city
    :param str source_state: name of source state
    :param str destination_city: name of destination city
    :param str destination_state: name of destination state
    """
    url = f'https://www.distance-cities.com/searchbd' \
          f'?from={source_city}%2C{source_state}' \
          f'&to={destination_city}%2C{destination_state}'
    url = url.replace(' ', '%2C')

    islands = ['Puerto Rico', 'Hawaii']
    if source_state in islands or destination_state in islands:
        logging.info(msg=f'{source_city}, {source_state} '
                         f'--> {destination_city}, {destination_state}: '
                         f'\t{url}')

    try:
        headers = {'Accept':     'text/html',
                   "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) "
                                 "AppleWebKit/537.36 (KHTML, like Gecko) "
                                 "Chrome/72.0.3626.121 Safari/537.36"}

        res = requests.get(url=url, headers=headers)
        res.raise_for_status()

        soup = BeautifulSoup(res.content, features='lxml')
        distance_elem = soup.select('#rkm')

        if distance_elem:
            value = float(distance_elem[0].text.split()[0].replace(',', ''))
            return value
        # if #rkm isn't found, must retrieve straight distance (no driving distance available)
        # this requires retrieval of JS vars, so requests_html is used
        else:
            session = HTMLSession()
            r = session.get(url)
            r.html.render(retries=50, timeout=120)
            straight_distance = r.html.find('#straightkm', first=True)
            logging.info(msg=f'{source_city}, {source_state} '
                             f'--> {destination_city}, {destination_state}: '
                             f'\t{straight_distance.text} km')
            if straight_distance.text:
                return float(straight_distance.text.replace(',', ''))

        logging.warning(msg=f'failed {source_city}, {source_state} '
                            f'--> {destination_city}, {destination_state}'.upper())
    except requests.exceptions.HTTPError as e:
        print(f'Error downloading webpage {url}', e)
        logging.warning(msg=f'Error downloading webpage {url}')
        logging.exception(msg=e)


def logger_config():
    """
    Sets up root logger and disables propagation from dependencies
    """
    logging.basicConfig(filename='scrape_distances.log',
                        level=logging.INFO,
                        format=' %(asctime)s.%(msecs)03d - %(levelname)s - '
                               '<%(funcName)s>: %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')
    logging.getLogger('requests').propagate = False
    logging.getLogger("urllib3").propagate = False
    logging.getLogger("pyppeteer").setLevel(logging.ERROR)


def patch_pyppeteer():
    """
    Disables pyppeteer's timeout (terminates at 20 seconds otherwise)
    """
    import pyppeteer.connection
    original_method = pyppeteer.connection.websockets.client.connect

    def new_method(*args, **kwargs):
        kwargs['ping_interval'] = None
        kwargs['ping_timeout'] = None
        return original_method(*args, **kwargs)

    pyppeteer.connection.websockets.client.connect = new_method


if __name__ == '__main__':
    logger_config()
    patch_pyppeteer()

    neighbor_regions = {1:  [9],
                        2:  [9, 10, 11],
                        3:  [4, 8, 11],  # 3 -> 8?
                        4:  [3, 5, 8],
                        5:  [4, 6, 8],
                        6:  [5, 7, 8],
                        7:  [6, 8, 10],  # 7 -> 11?
                        8:  [3, 4, 5, 6, 7],  # 8 -> 11?
                        9:  [1, 2],
                        10: [2, 7, 11],
                        11: [2, 3, 10]}  # 11 -> 7/8?

    root = join(abspath('.'), 'export', 'shelve')
    db = shelve.open(join(root, 'distance_vector'))

    path = join(abspath('.'), 'import', 'workbooks',
                'National_Transplant_Hospitals.xlsx')
    workbook = openpyxl.load_workbook(filename=path)
    sheet = workbook.active

    column_indices = set_default_indices()
    get_column_indices(worksheet=sheet, columns=column_indices)

    print('\n'.join(map(str, column_indices.items())))

    # cities = get_unique_locations(worksheet=sheet)
    # distance_matrix = set_default_distances(locations=cities)
    # distance_matrix = get_distances(distance_vector=distance_matrix)
    # db['distance_vector2'] = distance_matrix

    distance_matrix = db['distance_vector2']

    hospital_network = import_nodes(worksheet=sheet)
    db['hospital_network2'] = hospital_network
    print(hospital_network)
