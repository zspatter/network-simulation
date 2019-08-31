from pathlib import Path
import shelve

import openpyxl


def get_unique_states(worksheet):
    column = 'A'
    states = set()

    for x in range(2, worksheet.max_row + 1):
        states.add(worksheet[f'{column}{x}'].value)

    return sorted(states)


def get_unique_regions(worksheet):
    column = 'E'
    regions = set()

    for x in range(2, worksheet.max_row + 1):
        regions.add(int(worksheet[f'{column}{x}'].value))

    return sorted(regions)


def set_default_values(collection):
    collection_dict = {}

    for x in collection:
        collection_dict[x] = 0

    return collection_dict


def quantify_by_state(worksheet, state_dict):
    column = 'A'

    for x in range(2, worksheet.max_row + 1):
        state = worksheet[f'{column}{x}'].value
        state_dict[state] += 1

    return state_dict


def quantify_by_region(worksheet, region_dict):
    column = 'E'

    for x in range(2, worksheet.max_row + 1):
        region = int(worksheet[f'{column}{x}'].value)
        region_dict[region] += 1

    return region_dict


if __name__ == '__main__':
    path = Path('./import/workbooks/National_Transplant_Hospitals.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    unique_states = get_unique_states(sheet)
    unique_regions = get_unique_regions(sheet)

    states = set_default_values(unique_states)
    regions = set_default_values(unique_regions)

    states = quantify_by_state(sheet, states)
    regions = quantify_by_region(sheet, regions)

    export_path = Path('./export/shelve/hospital_quantities')
    db = shelve.open(str(export_path))

    db['states'] = states
    db['regions'] = regions
