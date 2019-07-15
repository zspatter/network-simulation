import json
from os.path import abspath, join

import openpyxl
import requests


def set_default_indices():
    """
    Sets default dict for columns (placeholders for indices)
    """
    fields = ('unique id',
              'hospital name',
              'city',
              'state',
              'region',
              'latitude',
              'longitude')
    columns = {}
    for field in fields:
        columns.setdefault(field, None)

    return columns


def get_column_indices(worksheet, columns):
    """
    Gets index for each column for reading data

    :param Worksheet worksheet: worksheet to read
    :param dict columns: expected columns to store indices
    """
    for x in range(1, 9):
        cell = worksheet.cell(row=1, column=x).value.lower()
        if cell in columns:
            columns[cell] = x


def get_unique_locations(worksheet):
    # creates a set of unique locations
    locations = set()
    for x in range(2, worksheet.max_row + 1):
        locations.add((worksheet.cell(row=x, column=column_indices['city']).value,
                       worksheet.cell(row=x, column=column_indices['state']).value,
                       int(worksheet.cell(row=x, column=column_indices['region']).value)))
    return locations


def get_coordinates(locations):
    location_dict = dict()

    for location in locations:
        city, state, _ = location
        coordinates = get_coordinate(city=city, state=state)
        if coordinates:
            location_dict[location] = coordinates

    return location_dict


def get_coordinate(city, state):
    url = f'https://dev.virtualearth.net/REST/v1/Locations/{city}%20{state}?&key={API_KEY}'

    try:
        response = requests.get(url=url)
        response.raise_for_status()

        location_data = json.loads(response.text)
        return location_data['resourceSets'][0]['resources'][0]['point']['coordinates']
    except requests.exceptions.HTTPError:
        print(city, state)


def set_coordinates(worksheet, locations):
    for x in range(2, worksheet.max_row + 1):
        city = worksheet.cell(row=x, column=column_indices['city']).value
        state = worksheet.cell(row=x, column=column_indices['state']).value
        region = int(worksheet.cell(row=x, column=column_indices['region']).value)

        if (city, state, region) in locations:
            lat, long = locations[(city, state, region)]
            worksheet.cell(row=x, column=column_indices['latitude']).value = lat
            worksheet.cell(row=x, column=column_indices['longitude']).value = long


if __name__ == '__main__':
    API_KEY = 'AmgxgtYUtWmbo4BUN5PraPs0T5sV-o5oUkJN74PsCwg3-BxE-DgOgZFAaQH1wIzx'
    path = join(abspath('.'), 'import', 'workbooks', 'National_Transplant_Hospitals.xlsx')
    workbook = openpyxl.load_workbook(filename=path)
    sheet = workbook.active

    column_indices = set_default_indices()
    get_column_indices(worksheet=sheet, columns=column_indices)

    unique_locations = get_unique_locations(worksheet=sheet)
    location_details = get_coordinates(locations=unique_locations)

    set_coordinates(worksheet=sheet, locations=location_details)
    workbook.save(path[:-5] + '_coordinates.xlsx')
